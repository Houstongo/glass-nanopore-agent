from __future__ import annotations

import argparse
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook


RESULT_ROOT = Path(r"D:\LabOSData\cvdata")
DEFAULT_DB_PATH = Path(r"D:\AntigravityProject\data\legacy_etching_import.sqlite")

IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace(" ", "")
    text = text.replace("\n", "")
    text = text.replace("（", "(").replace("）", ")")
    return text


def cell_to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def cell_to_int(value: Any) -> int | None:
    number = cell_to_float(value)
    if number is None:
        return None
    return int(number)


def determine_algorithm(path: Path) -> str:
    lower = str(path).lower()
    if "noise" in lower:
        return "FNA-MOBO"
    if "mixedcv" in lower or "4obj" in lower:
        return "IST-MOBO"
    return "Unknown"


def determine_target_angle(path: Path) -> int | None:
    matches = re.findall(r"(?<!\d)(15|20|25|30)(?!\d)", str(path))
    if matches:
        return int(matches[-1])
    degree_match = re.search(r"(\d+)\s*°", path.stem)
    if degree_match:
        return int(degree_match.group(1))
    return None


def normalize_stage_name(stem: str) -> str:
    stem = stem.strip()
    if stem.endswith("_input"):
        stem = stem[: -len("_input")]
    if stem.endswith("_result"):
        stem = stem[: -len("_result")]
    if stem.endswith("_summary"):
        stem = stem[: -len("_summary")]
    if stem == "input":
        return "initial"
    if stem == "group":
        return "unknown"
    if stem == "initial":
        return "initial"
    match = re.match(r"initial\d+", stem)
    if match:
        return match.group(0)
    return stem


def parse_group_label(label: Any, default_stage: str, default_target: int | None) -> tuple[str, int | None, int | None]:
    if isinstance(label, (int, float)) and label == label:
        return default_stage, int(label), default_target

    text = normalize_text(label)
    if not text:
        return default_stage, None, default_target

    if text.isdigit():
        return default_stage, int(text), default_target

    match = re.match(r"^(?:(\d+)_)?([A-Za-z]+\d+|initial\d+|opt\d+|initial)-(\d+)$", text)
    if match:
        target = int(match.group(1)) if match.group(1) else default_target
        stage = match.group(2)
        group_index = int(match.group(3))
        return stage, group_index, target

    match = re.match(r"^([A-Za-z]+\d+|initial\d+|opt\d+|initial)-(\d+)$", text)
    if match:
        return match.group(1), int(match.group(2)), default_target

    return default_stage, None, default_target


def parse_image_record_label(label: Any) -> tuple[int | None, int | None]:
    text = normalize_text(label)
    if not text:
        return None, None

    parts = text.split("-")
    if len(parts) >= 2 and parts[-1].isdigit() and parts[-2].isdigit():
        return int(parts[-2]), int(parts[-1])

    if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
        return int(parts[0]), int(parts[1])

    return None, None


def find_header_row(worksheet, required_headers: set[str]) -> tuple[int, dict[str, int]] | tuple[None, None]:
    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        columns: dict[str, int] = {}
        seen = {normalize_text(value): idx for idx, value in enumerate(row) if normalize_text(value)}
        if required_headers.issubset(seen.keys()):
            for name in required_headers:
                columns[name] = seen[name]
            return row_index, columns
    return None, None


def safe_load_workbook(workbook_path: Path):
    try:
        return load_workbook(workbook_path, data_only=True)
    except (BadZipFile, OSError, ValueError):
        return None


@dataclass
class ExperimentRow:
    dataset_root: str
    source_workbook: str
    source_sheet: str
    stage_name: str
    group_name: str
    group_index: int | None
    algorithm: str
    target_cone_angle_deg: int | None
    cone_angle_deg: float | None = None
    angle_diff_deg: float | None = None
    positive_voltage_v: float | None = None
    negative_voltage_v: float | None = None
    immersion_depth_um: float | None = None
    frequency_hz: float | None = None
    stability: float | None = None
    angle_variance: float | None = None
    notes: str | None = None


class ExperimentCollector:
    def __init__(self) -> None:
        self.rows: dict[tuple[str, str, int | None, int | None], ExperimentRow] = {}
        self.images: list[dict[str, Any]] = []

    def get_or_create(
        self,
        dataset_root: Path,
        workbook_path: Path,
        source_sheet: str,
        stage_name: str,
        group_index: int | None,
        target_cone_angle_deg: int | None,
    ) -> ExperimentRow:
        key = (str(dataset_root), stage_name, group_index, target_cone_angle_deg)
        if key not in self.rows:
            group_name = f"{stage_name}-{group_index}" if group_index is not None else stage_name
            self.rows[key] = ExperimentRow(
                dataset_root=str(dataset_root),
                source_workbook=str(workbook_path),
                source_sheet=source_sheet,
                stage_name=stage_name,
                group_name=group_name,
                group_index=group_index,
                algorithm=determine_algorithm(dataset_root),
                target_cone_angle_deg=target_cone_angle_deg,
            )
        return self.rows[key]


def import_parameter_workbook(collector: ExperimentCollector, dataset_root: Path, workbook_path: Path) -> None:
    workbook = safe_load_workbook(workbook_path)
    if workbook is None:
        return
    required = {"正电压(V)", "负电压(V)", "频率(Hz)", "伸入长度(μm)"}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row, columns = find_header_row(worksheet, required)
        if header_row is None:
            continue

        stage_name = normalize_stage_name(workbook_path.stem)
        default_target = determine_target_angle(workbook_path)
        seen_rows = 0
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            values = [row[idx] if idx < len(row) else None for idx in range(max(columns.values()) + 1)]
            if all(value is None for value in values):
                continue

            seen_rows += 1
            group_col = None
            optional_seen = {normalize_text(cell): idx for idx, cell in enumerate(worksheet[header_row]) if normalize_text(cell)}
            if "组号" in optional_seen:
                group_col = optional_seen["组号"]

            group_index = cell_to_int(row[group_col]) if group_col is not None else seen_rows
            experiment = collector.get_or_create(
                dataset_root=dataset_root,
                workbook_path=workbook_path,
                source_sheet=sheet_name,
                stage_name=stage_name,
                group_index=group_index,
                target_cone_angle_deg=default_target,
            )
            experiment.positive_voltage_v = cell_to_float(row[columns["正电压(V)"]])
            experiment.negative_voltage_v = cell_to_float(row[columns["负电压(V)"]])
            experiment.frequency_hz = cell_to_float(row[columns["频率(Hz)"]])
            experiment.immersion_depth_um = cell_to_float(row[columns["伸入长度(μm)"]])

            header_map = {normalize_text(cell.value): idx for idx, cell in enumerate(worksheet[header_row]) if normalize_text(cell.value)}
            angle_col = header_map.get("角度差值") or header_map.get("角度差值_mean")
            stability_col = header_map.get("稳定性")
            variance_col = header_map.get("角度差值方差") or header_map.get("角度差值_var")
            if angle_col is not None:
                experiment.angle_diff_deg = cell_to_float(row[angle_col])
            if stability_col is not None:
                experiment.stability = cell_to_float(row[stability_col])
            if variance_col is not None:
                experiment.angle_variance = cell_to_float(row[variance_col])


def import_summary_workbook(collector: ExperimentCollector, dataset_root: Path, workbook_path: Path) -> None:
    workbook = safe_load_workbook(workbook_path)
    if workbook is None:
        return
    required = {"组号"}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row, columns = find_header_row(worksheet, required)
        if header_row is None:
            continue

        header_map = {normalize_text(cell.value): idx for idx, cell in enumerate(worksheet[header_row]) if normalize_text(cell.value)}
        default_stage = normalize_stage_name(workbook_path.stem)
        default_target = determine_target_angle(workbook_path)

        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(cell is not None for cell in row):
                continue

            stage_name, group_index, target_angle = parse_group_label(
                row[columns["组号"]],
                default_stage=default_stage,
                default_target=default_target,
            )
            experiment = collector.get_or_create(
                dataset_root=dataset_root,
                workbook_path=workbook_path,
                source_sheet=sheet_name,
                stage_name=stage_name,
                group_index=group_index,
                target_cone_angle_deg=target_angle,
            )

            angle_col = header_map.get("角度_mean")
            diff_col = header_map.get("角度差值") or header_map.get("角度差值_mean")
            stability_col = header_map.get("稳定性")
            variance_col = header_map.get("角度差值方差") or header_map.get("角度差值_var")

            if angle_col is not None:
                experiment.cone_angle_deg = cell_to_float(row[angle_col])
            if diff_col is not None:
                experiment.angle_diff_deg = cell_to_float(row[diff_col])
            if stability_col is not None:
                experiment.stability = cell_to_float(row[stability_col])
            if variance_col is not None:
                experiment.angle_variance = cell_to_float(row[variance_col])


def pick_stage_directory(dataset_root: Path, stage_name: str) -> Path | None:
    exact = dataset_root / stage_name
    if exact.is_dir():
        return exact

    if stage_name in {"initial", "initial10"}:
        candidates = sorted([path for path in dataset_root.iterdir() if path.is_dir() and path.name.startswith("initial")])
        if len(candidates) == 1:
            return candidates[0]

    if stage_name.startswith("initial"):
        candidates = sorted([path for path in dataset_root.iterdir() if path.is_dir() and path.name.startswith("initial")])
        if candidates:
            return candidates[0]

    return None


def load_group_image_metrics(metrics_path: Path) -> dict[int, dict[str, float | None]]:
    if not metrics_path.exists():
        return {}

    workbook = safe_load_workbook(metrics_path)
    if workbook is None:
        return {}
    worksheet = workbook[workbook.sheetnames[0]]
    header_row, columns = find_header_row(worksheet, {"编号"})
    if header_row is None:
        return {}

    header_map = {normalize_text(cell.value): idx for idx, cell in enumerate(worksheet[header_row]) if normalize_text(cell.value)}
    result: dict[int, dict[str, float | None]] = {}
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        _, image_index = parse_image_record_label(row[columns["编号"]])
        if image_index is None:
            continue
        result[image_index] = {
            "measured_cone_angle_deg": cell_to_float(row[header_map["角度"]]) if "角度" in header_map else None,
            "measured_angle_diff_deg": cell_to_float(row[header_map["角度差值"]]) if "角度差值" in header_map else None,
        }
    return result


def map_images(collector: ExperimentCollector) -> None:
    for experiment in collector.rows.values():
        if experiment.group_index is None:
            continue

        dataset_root = Path(experiment.dataset_root)
        stage_dir = pick_stage_directory(dataset_root, experiment.stage_name)
        if stage_dir is None:
            continue

        group_dir = stage_dir / str(experiment.group_index)
        if not group_dir.is_dir():
            continue

        metrics = load_group_image_metrics(group_dir / "output_results.xlsx")
        image_files = sorted([path for path in group_dir.iterdir() if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES])
        for image_path in image_files:
            image_index_match = re.match(r"(\d+)", image_path.stem)
            image_index = int(image_index_match.group(1)) if image_index_match else None
            metric = metrics.get(image_index or -1, {})
            collector.images.append(
                {
                    "dataset_root": experiment.dataset_root,
                    "stage_name": experiment.stage_name,
                    "group_index": experiment.group_index,
                    "target_cone_angle_deg": experiment.target_cone_angle_deg,
                    "image_path": str(image_path),
                    "image_name": image_path.name,
                    "image_index": image_index,
                    "metrics_workbook": str(group_dir / "output_results.xlsx") if (group_dir / "output_results.xlsx").exists() else None,
                    "measured_cone_angle_deg": metric.get("measured_cone_angle_deg"),
                    "measured_angle_diff_deg": metric.get("measured_angle_diff_deg"),
                }
            )


def create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        DROP TABLE IF EXISTS experiment_images;
        DROP TABLE IF EXISTS experiments;

        CREATE TABLE experiments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset_root TEXT NOT NULL,
            source_workbook TEXT NOT NULL,
            source_sheet TEXT NOT NULL,
            stage_name TEXT NOT NULL,
            group_name TEXT NOT NULL,
            group_index INTEGER,
            algorithm TEXT NOT NULL,
            target_cone_angle_deg INTEGER,
            cone_angle_deg REAL,
            angle_diff_deg REAL,
            positive_voltage_v REAL,
            negative_voltage_v REAL,
            immersion_depth_um REAL,
            frequency_hz REAL,
            stability REAL,
            angle_variance REAL,
            notes TEXT
        );

        CREATE TABLE experiment_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            experiment_id INTEGER NOT NULL,
            image_path TEXT NOT NULL,
            image_name TEXT NOT NULL,
            image_index INTEGER,
            metrics_workbook TEXT,
            measured_cone_angle_deg REAL,
            measured_angle_diff_deg REAL,
            FOREIGN KEY (experiment_id) REFERENCES experiments(id)
        );
        """
    )


def write_database(db_path: Path, collector: ExperimentCollector) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(db_path)
    create_schema(connection)

    experiment_ids: dict[tuple[str, str, int | None, int | None], int] = {}
    for key, experiment in sorted(collector.rows.items()):
        cursor = connection.execute(
            """
            INSERT INTO experiments (
                dataset_root, source_workbook, source_sheet, stage_name, group_name, group_index,
                algorithm, target_cone_angle_deg, cone_angle_deg, angle_diff_deg,
                positive_voltage_v, negative_voltage_v, immersion_depth_um, frequency_hz,
                stability, angle_variance, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                experiment.dataset_root,
                experiment.source_workbook,
                experiment.source_sheet,
                experiment.stage_name,
                experiment.group_name,
                experiment.group_index,
                experiment.algorithm,
                experiment.target_cone_angle_deg,
                experiment.cone_angle_deg,
                experiment.angle_diff_deg,
                experiment.positive_voltage_v,
                experiment.negative_voltage_v,
                experiment.immersion_depth_um,
                experiment.frequency_hz,
                experiment.stability,
                experiment.angle_variance,
                experiment.notes,
            ),
        )
        experiment_ids[key] = int(cursor.lastrowid)

    for image in collector.images:
        key = (
            image["dataset_root"],
            image["stage_name"],
            image["group_index"],
            image["target_cone_angle_deg"],
        )
        experiment_id = experiment_ids.get(key)
        if experiment_id is None:
            continue
        connection.execute(
            """
            INSERT INTO experiment_images (
                experiment_id, image_path, image_name, image_index, metrics_workbook,
                measured_cone_angle_deg, measured_angle_diff_deg
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                experiment_id,
                image["image_path"],
                image["image_name"],
                image["image_index"],
                image["metrics_workbook"],
                image["measured_cone_angle_deg"],
                image["measured_angle_diff_deg"],
            ),
        )

    connection.commit()
    connection.close()


def collect_experiments(result_root: Path) -> ExperimentCollector:
    collector = ExperimentCollector()
    for dataset_root in sorted(path for path in result_root.iterdir() if path.is_dir()):
        root_workbooks = sorted(dataset_root.glob("*.xlsx"))
        if not root_workbooks:
            continue

        for workbook_path in root_workbooks:
            stem = workbook_path.stem.lower()
            if stem.endswith("_summary") or stem == "group":
                import_summary_workbook(collector, dataset_root, workbook_path)
            elif stem.endswith("_result"):
                import_summary_workbook(collector, dataset_root, workbook_path)
            else:
                import_parameter_workbook(collector, dataset_root, workbook_path)
                if stem.endswith(".xlsx"):
                    pass

        # Older merged files such as 15°.xlsx contain both parameters and outputs in one sheet.
        for workbook_path in root_workbooks:
            if re.search(r"\d+\s*°", workbook_path.stem):
                import_parameter_workbook(collector, dataset_root, workbook_path)

    map_images(collector)
    return collector


def build_database(result_root: Path, db_path: Path) -> tuple[int, int]:
    collector = collect_experiments(result_root)
    write_database(db_path, collector)
    return len(collector.rows), len(collector.images)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a SQLite database for Yang Xuelian experiment data.")
    parser.add_argument("--result-root", type=Path, default=RESULT_ROOT)
    parser.add_argument("--db-path", type=Path, default=DEFAULT_DB_PATH)
    args = parser.parse_args()

    experiment_count, image_count = build_database(args.result_root, args.db_path)
    print(f"result_root={args.result_root}")
    print(f"db_path={args.db_path}")
    print(f"experiments={experiment_count}")
    print(f"images={image_count}")


if __name__ == "__main__":
    main()
