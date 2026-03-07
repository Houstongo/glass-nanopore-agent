from __future__ import annotations

import argparse
import json
import re
import sqlite3
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


CV_RESULT_ROOT = Path(r"D:\LabOSData\result")
WORKBOOK_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}

PARAMETER_FIELD_ALIASES = {
    "正电压(v)": "positive_voltage_v",
    "负电压(v)": "negative_voltage_v",
    "频率(hz)": "frequency_hz",
    "伸入长度(μm)": "immersion_depth_um",
    "伸入长度(um)": "immersion_depth_um",
}

MEASUREMENT_FIELD_ALIASES = {
    "角度_mean": "cone_angle_deg",
    "角度": "cone_angle_deg",
    "角度差值_mean": "angle_diff_deg",
    "角度差值": "angle_diff_deg",
    "稳定性": "stability",
}

SINGLE_FACTOR_STAGE_PARAMETER_MAP = {
    "var1_solution_concentration": "solution_concentration",
    "var2_positive_voltage": "positive_voltage_v",
    "var3_negative_voltage": "negative_voltage_v",
    "var4_frequency": "frequency_hz",
    "var5_time": "etching_time_s",
    "var6_length": "immersion_depth_um",
}
DEFAULT_DB_PATH = Path(r"D:\AntigravityProject\data\etching_experiments.sqlite")
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "")
        .replace("\n", "")
        .replace("（", "(")
        .replace("）", ")")
        .replace("µ", "μ")
    )


def cell_to_float(value: Any) -> float | None:
    if value in {None, ""}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def cell_to_int(value: Any) -> int | None:
    number = cell_to_float(value)
    if number is None:
        return None
    return int(number)


def determine_algorithm(path: Path) -> str:
    lower = str(path).lower()
    if "noise" in lower:
        return "FNA-MOBO"
    if "mixedcv" in lower or "4obj" in lower:
        return "IST-MOBO"
    if "singlefactor" in lower:
        return "SingleFactor"
    return "Unknown"


def infer_dataset_family(path: Path) -> str:
    lower = str(path).lower()
    if "noise" in lower:
        return "noise"
    if "singlefactor" in lower:
        return "single_factor"
    if "mixedcv" in lower or "4obj" in lower:
        return "mixedcv"
    if "sem" in lower:
        return "sem"
    return "other"


def determine_target_angle(path: Path) -> int | None:
    matches = re.findall(r"(?<!\d)(15|20|25|30)(?!\d)", str(path))
    if matches:
        return int(matches[-1])
    degree_match = re.search(r"(\d+)\s*[°º掳]", path.stem)
    if degree_match:
        return int(degree_match.group(1))
    return None


def normalize_stage_name(stem: str) -> str:
    stem = stem.strip()
    for suffix in ("_input", "_result", "_summary"):
        if stem.endswith(suffix):
            stem = stem[: -len(suffix)]
    if stem in {"output", "output_result", "output_results"}:
        return "unknown"
    if stem == "input":
        return "initial"
    if stem == "group":
        return "unknown"
    return stem


def parse_group_label(
    label: Any, default_stage: str, default_target: int | None
) -> tuple[str, int | None, int | None]:
    if isinstance(label, (int, float)) and label == label:
        return default_stage, int(label), default_target

    text = normalize_text(label)
    if not text:
        return default_stage, None, default_target

    if text.isdigit():
        return default_stage, int(text), default_target

    match = re.match(r"^(?:(\d+)_)?([A-Za-z]+\d+|initial\d+|opt\d+|initial)-(\d+)$", text)
    if match:
        target = int(match.group(1)) if match.group(1) else default_target
        return match.group(2), int(match.group(3)), target

    return default_stage, None, default_target


def parse_image_record_label(label: Any) -> tuple[int | None, int | None]:
    text = normalize_text(label)
    if not text:
        return None, None

    parts = text.split("-")
    if len(parts) >= 2 and parts[-1].isdigit() and parts[-2].isdigit():
        return int(parts[-2]), int(parts[-1])
    return None, None


def parse_stage_and_group_from_directory(name: str) -> tuple[str, int | None]:
    text = name.strip()
    if text.isdigit():
        return "", int(text)

    match = re.match(r"^\d+_(initial\d+|opt\d+|initial)$", text)
    if match:
        return match.group(1), None

    match = re.match(r"^(initial\d+|opt\d+|initial)-(\d+)$", text)
    if match:
        return match.group(1), int(match.group(2))

    if re.match(r"^(initial\d+|opt\d+|initial)$", text):
        return text, None

    return text, None


def parse_image_path(image_path: Path, dataset_root: Path) -> dict[str, Any]:
    relative_parts = image_path.relative_to(dataset_root).parts
    if len(relative_parts) < 2:
        raise ValueError(f"图片路径层级不足，无法解析: {image_path}")

    image_index_match = re.match(r"(\d+)", image_path.stem)
    image_index = int(image_index_match.group(1)) if image_index_match else None

    stage_name: str | None = None
    group_index: int | None = None
    directory_parts = list(relative_parts[:-1])

    # 从右往左解析，优先用最接近图片文件的目录恢复组号，再向左找最近的阶段目录。
    for index in range(len(directory_parts) - 1, -1, -1):
        candidate_stage, candidate_group = parse_stage_and_group_from_directory(directory_parts[index])
        if candidate_group is not None:
            group_index = candidate_group
            if candidate_stage:
                stage_name = candidate_stage
                break
            for left_index in range(index - 1, -1, -1):
                left_stage, left_group = parse_stage_and_group_from_directory(directory_parts[left_index])
                if left_stage:
                    stage_name = left_stage
                    if group_index is None and left_group is not None:
                        group_index = left_group
                    break
            break

    if group_index is None:
        # 有些数据把组号直接编码进图片文件名，例如 opt4/1-15s未刻完.jpg。
        group_match = re.match(r"(\d+)", image_path.stem)
        group_index = int(group_match.group(1)) if group_match else None
        for index in range(len(directory_parts) - 1, -1, -1):
            candidate_stage, _ = parse_stage_and_group_from_directory(directory_parts[index])
            if candidate_stage:
                stage_name = candidate_stage
                break

    if group_index is None:
        # SEM 文件常见命名：opt2-2-1-4w.tiff，其中依次是阶段、组号、图片序号、倍率。
        normalized_stem = image_path.stem.strip().replace(" ", "")
        sem_match = re.match(r"^(initial\d+|opt\d+|initial)-(\d+)-(\d+)(?:-.+)?$", normalized_stem)
        if sem_match:
            stage_name = sem_match.group(1)
            group_index = int(sem_match.group(2))
            image_index = int(sem_match.group(3))

    if group_index is None:
        raise ValueError(f"无法从图片目录恢复实验组: {image_path}")
    if not stage_name:
        raise ValueError(f"无法从图片路径恢复实验阶段: {image_path}")

    return {
        "stage_name": stage_name,
        "group_index": group_index,
        "image_index": image_index,
        "image_name": image_path.name,
        "image_path": str(image_path),
        "capture_stage": stage_name,
        "image_role": "tip",
    }


@dataclass
class DatasetRecord:
    dataset_name: str
    dataset_family: str
    algorithm_name: str
    target_cone_angle_deg: int | None
    source_root_path: str
    description: str | None = None


@dataclass
class ExperimentRecord:
    dataset_name: str
    stage_name: str
    group_name: str
    group_index: int | None
    run_label: str | None = None
    sample_count: int = 0
    status: str = "active"
    notes: str | None = None
    parameters: dict[str, Any] = field(default_factory=dict)
    measurements: dict[str, Any] = field(default_factory=dict)


@dataclass
class ExperimentCollector:
    datasets: dict[str, DatasetRecord] = field(default_factory=dict)
    experiments: dict[tuple[str, str, int | None], ExperimentRecord] = field(default_factory=dict)
    images: list[dict[str, Any]] = field(default_factory=list)
    source_files: list[dict[str, Any]] = field(default_factory=list)
    raw_import_rows: list[dict[str, Any]] = field(default_factory=list)
    next_source_ref_id: int = 1

    def get_or_create_dataset(self, dataset_root: Path) -> DatasetRecord:
        dataset_name = dataset_root.name
        dataset = self.datasets.get(dataset_name)
        if dataset is None:
            dataset = DatasetRecord(
                dataset_name=dataset_name,
                dataset_family=infer_dataset_family(dataset_root),
                algorithm_name=determine_algorithm(dataset_root),
                target_cone_angle_deg=determine_target_angle(dataset_root),
                source_root_path=str(dataset_root),
            )
            self.datasets[dataset_name] = dataset
        return dataset

    def get_or_create_experiment(
        self, dataset_root: Path, stage_name: str, group_index: int | None
    ) -> ExperimentRecord:
        dataset = self.get_or_create_dataset(dataset_root)
        key = (dataset.dataset_name, stage_name, group_index)
        experiment = self.experiments.get(key)
        if experiment is None:
            group_name = f"{stage_name}-{group_index}" if group_index is not None else stage_name
            experiment = ExperimentRecord(
                dataset_name=dataset.dataset_name,
                stage_name=stage_name,
                group_name=group_name,
                group_index=group_index,
                run_label=group_name,
            )
            self.experiments[key] = experiment
        return experiment

    def register_source_file(
        self,
        *,
        dataset_name: str,
        experiment_key: tuple[str, str, int | None] | None,
        file_path: str,
        file_type: str,
        sheet_name: str | None,
        source_kind: str,
        notes: str | None,
    ) -> int:
        source_ref_id = self.next_source_ref_id
        self.next_source_ref_id += 1
        self.source_files.append(
            {
                "source_ref_id": source_ref_id,
                "dataset_name": dataset_name,
                "experiment_key": experiment_key,
                "file_path": file_path,
                "file_type": file_type,
                "sheet_name": sheet_name,
                "source_kind": source_kind,
                "notes": notes,
            }
        )
        return source_ref_id


def find_header_row(
    worksheet, required_headers: set[str]
) -> tuple[int | None, dict[str, int] | None]:
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1
    ):
        seen = {normalize_text(value): idx for idx, value in enumerate(row) if normalize_text(value)}
        if required_headers.issubset(seen.keys()):
            return row_index, {name: seen[name] for name in required_headers}
    return None, None


def find_header_row_with_any_alias(
    worksheet, aliases: dict[str, str]
) -> int | None:
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1
    ):
        seen = {normalize_text(value) for value in row if normalize_text(value)}
        if seen.intersection(aliases.keys()):
            return row_index
    return None


def safe_load_workbook(workbook_path: Path):
    if load_workbook is None:
        return None
    try:
        return load_workbook(workbook_path, data_only=True)
    except Exception:
        return None


def _json_safe_value(value: Any) -> Any:
    if isinstance(value, (str, int, float)) or value is None:
        return value
    return str(value)


def infer_stage_and_group_from_workbook_path(
    workbook_path: Path, dataset_root: Path
) -> tuple[str, int | None]:
    relative_parts = workbook_path.relative_to(dataset_root).parts[:-1]
    stage_name: str | None = None
    group_index: int | None = None

    for index in range(len(relative_parts) - 1, -1, -1):
        candidate_stage, candidate_group = parse_stage_and_group_from_directory(relative_parts[index])
        if candidate_group is not None and group_index is None:
            group_index = candidate_group
        if candidate_stage and candidate_stage not in {"excel", "SEM"} and stage_name is None:
            stage_name = candidate_stage
        if stage_name and group_index is not None:
            break

    if not stage_name:
        stage_name = normalize_stage_name(workbook_path.stem)

    if stage_name == "unknown":
        for index in range(len(relative_parts) - 1, -1, -1):
            candidate_stage, _ = parse_stage_and_group_from_directory(relative_parts[index])
            if candidate_stage and candidate_stage not in {"excel", "SEM"}:
                stage_name = candidate_stage
                break

    return stage_name, group_index


def infer_group_from_row(
    row: tuple[Any, ...],
    header_map: dict[str, int],
    row_number: int,
) -> int | None:
    if "组号" in header_map:
        _, group_index, _ = parse_group_label(row[header_map["组号"]], "unknown", None)
        if group_index is not None:
            return group_index

    if "编号" in header_map:
        sample_group, _ = parse_image_record_label(row[header_map["编号"]])
        if sample_group is not None:
            return sample_group

    return row_number


def apply_header_mappings(
    target: dict[str, Any],
    row: tuple[Any, ...],
    header_map: dict[str, int],
    aliases: dict[str, str],
) -> None:
    for header_name, field_name in aliases.items():
        if header_name not in header_map:
            continue
        value = cell_to_float(row[header_map[header_name]])
        if value is not None:
            target.setdefault(field_name, value)


def apply_single_factor_parameter_mapping(experiment: ExperimentRecord) -> None:
    parameter_field = SINGLE_FACTOR_STAGE_PARAMETER_MAP.get(experiment.stage_name.lower())
    if parameter_field is None or experiment.group_index is None:
        return
    experiment.parameters.setdefault(parameter_field, float(experiment.group_index))


def should_import_workbook(workbook_path: Path, dataset_root: Path) -> bool:
    if workbook_path.suffix.lower() not in WORKBOOK_SUFFIXES or workbook_path.name.startswith("~$"):
        return False

    stem_lower = workbook_path.stem.lower()
    name_lower = workbook_path.name.lower()
    relative_parts = workbook_path.relative_to(dataset_root).parts
    is_root_workbook = len(relative_parts) == 1

    if "副本" in workbook_path.stem:
        return False
    if dataset_root.name == "SingleFactor" and workbook_path.stem == "汇总":
        return False
    if stem_lower == "result":
        return False
    if is_root_workbook and stem_lower == "output_results":
        return False
    if is_root_workbook and stem_lower.endswith("_results"):
        return False

    return True


def import_workbook_rows(collector: ExperimentCollector, dataset_root: Path, workbook_path: Path) -> None:
    workbook = safe_load_workbook(workbook_path)
    if workbook is None:
        collector.register_source_file(
            dataset_name=dataset_root.name,
            experiment_key=None,
            file_path=str(workbook_path),
            file_type=workbook_path.suffix.lstrip(".").lower(),
            sheet_name=None,
            source_kind="workbook_unreadable",
            notes="工作簿无法读取，已跳过",
        )
        return

    default_stage, path_group_index = infer_stage_and_group_from_workbook_path(workbook_path, dataset_root)
    default_target = determine_target_angle(workbook_path) or determine_target_angle(dataset_root)

    for sheet_name in workbook.sheetnames:
        source_ref_id = collector.register_source_file(
            dataset_name=dataset_root.name,
            experiment_key=None,
            file_path=str(workbook_path),
            file_type=workbook_path.suffix.lstrip(".").lower(),
            sheet_name=sheet_name,
            source_kind="workbook",
            notes=None,
        )
        worksheet = workbook[sheet_name]
        header_row, _ = find_header_row(worksheet, {"组号"})
        if header_row is None:
            header_row, _ = find_header_row(worksheet, {"编号"})
        if header_row is None:
            header_row = find_header_row_with_any_alias(worksheet, PARAMETER_FIELD_ALIASES)
        if header_row is None:
            header_row = find_header_row_with_any_alias(worksheet, MEASUREMENT_FIELD_ALIASES)
        if header_row is None:
            continue

        header_map = {
            normalize_text(cell.value): idx
            for idx, cell in enumerate(worksheet[header_row])
            if normalize_text(cell.value)
        }
        for row_offset, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=1,
        ):
            if not any(cell is not None for cell in row):
                continue

            stage_name = default_stage
            group_index = path_group_index or infer_group_from_row(row, header_map, row_offset)
            target_angle = default_target

            if "组号" in header_map:
                stage_name, parsed_group, target_angle = parse_group_label(
                    row[header_map["组号"]],
                    default_stage=default_stage,
                    default_target=default_target,
                )
                if parsed_group is not None:
                    group_index = parsed_group

            if group_index is None:
                continue

            experiment = collector.get_or_create_experiment(dataset_root, stage_name, group_index)
            if target_angle is not None:
                experiment.measurements.setdefault("target_cone_angle_deg", target_angle)
            apply_header_mappings(experiment.parameters, row, header_map, PARAMETER_FIELD_ALIASES)
            apply_header_mappings(experiment.measurements, row, header_map, MEASUREMENT_FIELD_ALIASES)
            apply_single_factor_parameter_mapping(experiment)

            collector.raw_import_rows.append(
                {
                    "source_ref_id": source_ref_id,
                    "dataset_name": dataset_root.name,
                    "experiment_key": (dataset_root.name, stage_name, group_index),
                    "row_identifier": f"{sheet_name}:{group_index}:{row_offset}",
                    "raw_json": json.dumps([_json_safe_value(cell) for cell in row], ensure_ascii=False),
                    "normalized_ok": 1,
                    "error_message": None,
                }
            )


def collect_experiments(result_root: Path) -> ExperimentCollector:
    collector = ExperimentCollector()
    if not result_root.exists():
        raise FileNotFoundError(f"data root does not exist: {result_root}")

    for dataset_root in sorted(
        path for path in result_root.iterdir() if path.is_dir() and path.name.lower() != "excel"
    ):
        collector.get_or_create_dataset(dataset_root)

        for workbook_path in sorted(
            path
            for path in dataset_root.rglob("*")
            if path.is_file() and should_import_workbook(path, dataset_root)
        ):
            import_workbook_rows(collector, dataset_root, workbook_path)

        for image_path in sorted(
            path
            for path in dataset_root.rglob("*")
            if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES
        ):
            parsed = parse_image_path(image_path, dataset_root)
            experiment = collector.get_or_create_experiment(
                dataset_root, parsed["stage_name"], parsed["group_index"]
            )
            experiment.sample_count += 1
            collector.images.append(
                {
                    "dataset_name": dataset_root.name,
                    "stage_name": parsed["stage_name"],
                    "group_index": parsed["group_index"],
                    "image_path": parsed["image_path"],
                    "image_name": parsed["image_name"],
                    "image_index": parsed["image_index"],
                    "image_role": parsed["image_role"],
                    "capture_stage": parsed["capture_stage"],
                    "measured_cone_angle_deg": None,
                    "measured_angle_diff_deg": None,
                    "metadata_json": None,
                }
            )
            collector.register_source_file(
                dataset_name=dataset_root.name,
                experiment_key=(dataset_root.name, parsed["stage_name"], parsed["group_index"]),
                file_path=parsed["image_path"],
                file_type=image_path.suffix.lstrip(".").lower(),
                sheet_name=None,
                source_kind="image",
                notes=None,
            )

    return collector


def create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        PRAGMA foreign_keys = ON;

        DROP TABLE IF EXISTS raw_import_rows;
        DROP TABLE IF EXISTS source_files;
        DROP TABLE IF EXISTS experiment_images;
        DROP TABLE IF EXISTS experiment_measurements;
        DROP TABLE IF EXISTS experiment_parameters;
        DROP TABLE IF EXISTS experiments;
        DROP TABLE IF EXISTS datasets;

        CREATE TABLE datasets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset_name TEXT NOT NULL UNIQUE,
            dataset_family TEXT NOT NULL,
            algorithm_name TEXT NOT NULL,
            target_cone_angle_deg INTEGER,
            source_root_path TEXT NOT NULL,
            description TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE experiments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset_id INTEGER NOT NULL,
            stage_name TEXT NOT NULL,
            group_name TEXT NOT NULL,
            group_index INTEGER,
            run_label TEXT,
            sample_count INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'active',
            notes TEXT,
            FOREIGN KEY (dataset_id) REFERENCES datasets(id),
            UNIQUE (dataset_id, stage_name, group_index)
        );

        CREATE TABLE experiment_parameters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            experiment_id INTEGER NOT NULL UNIQUE,
            positive_voltage_v REAL,
            negative_voltage_v REAL,
            frequency_hz REAL,
            immersion_depth_um REAL,
            solution_concentration REAL,
            etching_time_s REAL,
            tip_diameter_um REAL,
            capillary_diameter_um REAL,
            heating_count REAL,
            parameter_json TEXT,
            FOREIGN KEY (experiment_id) REFERENCES experiments(id)
        );

        CREATE TABLE experiment_measurements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            experiment_id INTEGER NOT NULL UNIQUE,
            cone_angle_deg REAL,
            target_cone_angle_deg REAL,
            angle_diff_deg REAL,
            stability REAL,
            angle_variance REAL,
            roughness REAL,
            symmetry_score REAL,
            quality_score REAL,
            measurement_json TEXT,
            FOREIGN KEY (experiment_id) REFERENCES experiments(id)
        );

        CREATE TABLE experiment_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            experiment_id INTEGER NOT NULL,
            image_path TEXT NOT NULL,
            image_name TEXT NOT NULL,
            image_index INTEGER,
            image_role TEXT,
            capture_stage TEXT,
            measured_cone_angle_deg REAL,
            measured_angle_diff_deg REAL,
            thumbnail_path TEXT,
            metadata_json TEXT,
            FOREIGN KEY (experiment_id) REFERENCES experiments(id)
        );

        CREATE TABLE source_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset_id INTEGER NOT NULL,
            experiment_id INTEGER,
            file_path TEXT NOT NULL,
            file_type TEXT NOT NULL,
            sheet_name TEXT,
            source_kind TEXT NOT NULL,
            checksum TEXT,
            notes TEXT,
            FOREIGN KEY (dataset_id) REFERENCES datasets(id),
            FOREIGN KEY (experiment_id) REFERENCES experiments(id)
        );

        CREATE TABLE raw_import_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file_id INTEGER,
            row_identifier TEXT,
            raw_json TEXT NOT NULL,
            normalized_ok INTEGER NOT NULL DEFAULT 1,
            error_message TEXT,
            FOREIGN KEY (source_file_id) REFERENCES source_files(id)
        );
        """
    )


def _ensure_parent_directory(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _json_or_none(payload: dict[str, Any]) -> str | None:
    cleaned = {key: value for key, value in payload.items() if value is not None}
    if not cleaned:
        return None
    return json.dumps(cleaned, ensure_ascii=False, sort_keys=True)


def write_database(db_path: Path, collector: ExperimentCollector) -> tuple[int, int, int]:
    _ensure_parent_directory(db_path)

    connection = sqlite3.connect(db_path)
    # 当前工作区磁盘环境下默认 DELETE journal 会触发 disk I/O error，
    # 改用 TRUNCATE 可稳定写入且仍保留回滚日志能力。
    connection.execute("PRAGMA journal_mode=TRUNCATE")
    create_schema(connection)

    dataset_ids: dict[str, int] = {}
    for dataset_name, dataset in sorted(collector.datasets.items()):
        cursor = connection.execute(
            """
            INSERT INTO datasets (
                dataset_name, dataset_family, algorithm_name, target_cone_angle_deg,
                source_root_path, description
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                dataset.dataset_name,
                dataset.dataset_family,
                dataset.algorithm_name,
                dataset.target_cone_angle_deg,
                dataset.source_root_path,
                dataset.description,
            ),
        )
        dataset_ids[dataset_name] = int(cursor.lastrowid)

    experiment_ids: dict[tuple[str, str, int | None], int] = {}
    for key, experiment in sorted(collector.experiments.items()):
        cursor = connection.execute(
            """
            INSERT INTO experiments (
                dataset_id, stage_name, group_name, group_index, run_label,
                sample_count, status, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                dataset_ids[experiment.dataset_name],
                experiment.stage_name,
                experiment.group_name,
                experiment.group_index,
                experiment.run_label,
                experiment.sample_count,
                experiment.status,
                experiment.notes,
            ),
        )
        experiment_id = int(cursor.lastrowid)
        experiment_ids[key] = experiment_id

        connection.execute(
            """
            INSERT INTO experiment_parameters (
                experiment_id, positive_voltage_v, negative_voltage_v, frequency_hz,
                immersion_depth_um, solution_concentration, etching_time_s,
                tip_diameter_um, capillary_diameter_um, heating_count, parameter_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                experiment_id,
                experiment.parameters.get("positive_voltage_v"),
                experiment.parameters.get("negative_voltage_v"),
                experiment.parameters.get("frequency_hz"),
                experiment.parameters.get("immersion_depth_um"),
                experiment.parameters.get("solution_concentration"),
                experiment.parameters.get("etching_time_s"),
                experiment.parameters.get("tip_diameter_um"),
                experiment.parameters.get("capillary_diameter_um"),
                experiment.parameters.get("heating_count"),
                _json_or_none(experiment.parameters),
            ),
        )
        connection.execute(
            """
            INSERT INTO experiment_measurements (
                experiment_id, cone_angle_deg, target_cone_angle_deg, angle_diff_deg,
                stability, angle_variance, roughness, symmetry_score, quality_score,
                measurement_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                experiment_id,
                experiment.measurements.get("cone_angle_deg"),
                experiment.measurements.get("target_cone_angle_deg")
                or collector.datasets[experiment.dataset_name].target_cone_angle_deg,
                experiment.measurements.get("angle_diff_deg"),
                experiment.measurements.get("stability"),
                experiment.measurements.get("angle_variance"),
                experiment.measurements.get("roughness"),
                experiment.measurements.get("symmetry_score"),
                experiment.measurements.get("quality_score"),
                _json_or_none(experiment.measurements),
            ),
        )

    source_file_ids: dict[int, int] = {}
    for source_file in collector.source_files:
        experiment_id = None
        if source_file["experiment_key"] is not None:
            experiment_id = experiment_ids.get(source_file["experiment_key"])
        cursor = connection.execute(
            """
            INSERT INTO source_files (
                dataset_id, experiment_id, file_path, file_type, sheet_name,
                source_kind, checksum, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                dataset_ids[source_file["dataset_name"]],
                experiment_id,
                source_file["file_path"],
                source_file["file_type"],
                source_file["sheet_name"],
                source_file["source_kind"],
                None,
                source_file["notes"],
            ),
        )
        source_file_ids[source_file["source_ref_id"]] = int(cursor.lastrowid)

    for image in collector.images:
        experiment_key = (image["dataset_name"], image["stage_name"], image["group_index"])
        connection.execute(
            """
            INSERT INTO experiment_images (
                experiment_id, image_path, image_name, image_index, image_role,
                capture_stage, measured_cone_angle_deg, measured_angle_diff_deg,
                thumbnail_path, metadata_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                experiment_ids[experiment_key],
                image["image_path"],
                image["image_name"],
                image["image_index"],
                image["image_role"],
                image["capture_stage"],
                image["measured_cone_angle_deg"],
                image["measured_angle_diff_deg"],
                None,
                image["metadata_json"],
            ),
        )

    for row in collector.raw_import_rows:
        source_file_id = source_file_ids.get(row.get("source_ref_id"))
        connection.execute(
            """
            INSERT INTO raw_import_rows (
                source_file_id, row_identifier, raw_json, normalized_ok, error_message
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (
                source_file_id,
                row["row_identifier"],
                row["raw_json"],
                row["normalized_ok"],
                row["error_message"],
            ),
        )

    connection.commit()
    connection.close()
    return len(collector.datasets), len(collector.experiments), len(collector.images)


def build_database(result_root: Path, db_path: Path) -> tuple[int, int, int]:
    collector = collect_experiments(result_root)
    return write_database(db_path, collector)


def main() -> None:
    parser = argparse.ArgumentParser(description="构建电刻蚀实验数据 SQLite 数据库。")
    parser.add_argument("--result-root", type=Path, default=CV_RESULT_ROOT)
    parser.add_argument("--db-path", type=Path, default=DEFAULT_DB_PATH)
    args = parser.parse_args()

    dataset_count, experiment_count, image_count = build_database(args.result_root, args.db_path)

    print(f"result_root={args.result_root}")
    print(f"db_path={args.db_path}")
    print(f"datasets={dataset_count}")
    print(f"experiments={experiment_count}")
    print(f"images={image_count}")


if __name__ == "__main__":
    main()
